from copy import deepcopy
from pathlib import Path
from io import BytesIO
from datetime import date
import json

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    REPORTLAB_DISPONIVEL = True
except ImportError:
    REPORTLAB_DISPONIVEL = False

from jirau_core import (
    CAMPOS_AREA, CAMPOS_INDICE, CAMPOS_VOLUME, DEFAULT_DB, SERVICOS_PROPOSTA,
    consumo_total_item, detectar_materiais_excel, encontrar_indice, encontrar_material,
    moeda, numero_br, numero_float, recalcular_teto, substituir_materiais_preservando_vinculos, uid,
    carregar_indice_tipo_oficial, auditar_indice,
)
from jirau_core import carregar_db as _carregar_db, salvar_db as _salvar_db, recalcular_todos as _recalcular_todos, migrar_json_para_sqlite
from premium_theme import premium_css, svg_data_uri

st.set_page_config(page_title="Jirau Enterprise | Orçamentos", page_icon="🏗️", layout="wide", initial_sidebar_state="expanded", menu_items={"About": "Jirau Enterprise V12.3 — Cálculos validados pela planilha oficial"})
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)
DB_FILE = DATA_DIR / "jirau_v11.sqlite3"  # Mantido para preservar os dados existentes
LEGACY_DB_FILE = DATA_DIR / "jirau_db.json"
migrar_json_para_sqlite(LEGACY_DB_FILE, DB_FILE)

def carregar_db():
    return _carregar_db(DB_FILE)

def salvar_db(db):
    return _salvar_db(db, DB_FILE)

def recalcular_todos(db):
    return _recalcular_todos(db, DB_FILE)

db = carregar_db()

JIRAU_RED = "C91F2B"
JIRAU_NAVY = "182733"
JIRAU_LIGHT = "E9EEF2"
JIRAU_GRAY = "D9D9D9"

def resumo_pavimentos(obra):
    return [{
        "PAVIMENTO": t.get("nome", ""), "ÁREA (m²)": float(t.get("area", 0)),
        "PÉ-DIREITO (m)": float(t.get("pe_direito", 0)), "VOLUME (m³)": float(t.get("volume", 0)),
        "DIAS": int(t.get("dias", 30)), "VALOR MENSAL": float(t.get("valor_mensal", 0)),
        "VALOR DO PERÍODO": float(t.get("valor_total", 0)),
    } for t in obra.get("tetos", [])]

def linhas_tecnicas(obra):
    linhas = []
    for teto in obra.get("tetos", []):
        for sv in teto.get("servicos", []):
            linhas.append({
                "QUANT. (JOGO)": sv.get("quantidade_jogos", 0), "DESCRIÇÃO": sv.get("descricao", ""),
                "%": sv.get("percentual", 0), "P.D. (m)": teto.get("pe_direito", 0),
                "PAVIMENTO": teto.get("nome", ""), "ÁREA m²": teto.get("area", 0),
                "m³": teto.get("volume", 0), "DIAS": teto.get("dias", 30),
                "UNIT. R$/DIA": sv.get("unitario_dia", 0), "TOTAL R$/MÊS": sv.get("total_mensal", 0),
                "TOTAL PERÍODO": sv.get("total_periodo", 0), "FATOR": sv.get("fator", 1),
                "HORIZONTAL R$/m²": sv.get("horizontal_rs_m2", 0), "VERTICAL R$/m³": sv.get("vertical_rs_m3", 0),
            })
    return linhas

def gerar_excel_proposta(obra):
    wb = Workbook()
    ws = wb.active
    ws.title = "Proposta Comercial"
    thin = Side(style="thin", color="1F1F1F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    red = PatternFill("solid", fgColor=JIRAU_RED)
    navy = PatternFill("solid", fgColor=JIRAU_NAVY)
    light = PatternFill("solid", fgColor=JIRAU_LIGHT)
    gray = PatternFill("solid", fgColor=JIRAU_GRAY)
    for i, width in enumerate([15, 38, 14, 14, 18, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.merge_cells("A1:F4")
    ws["A1"] = "ANDAIMES\nJIRAU\nESCORAMENTO"
    ws["A1"].fill = red
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for rr in range(1, 5): ws.row_dimensions[rr].height = 28
    ws.merge_cells("A6:F6")
    ws["A6"] = "PROPOSTA COMERCIAL"
    ws["A6"].font = Font(bold=True, size=18, color="FFFFFF")
    ws["A6"].fill = navy
    ws["A6"].alignment = Alignment(horizontal="center")
    dados = [("DATA", date.today().strftime("%d/%m/%Y")), ("CLIENTE", obra.get("cliente", "")), ("OBRA", obra.get("nome", "")), ("ENDEREÇO", obra.get("endereco", "")), ("REFERÊNCIA", obra.get("referencia", ""))]
    r = 8
    for rotulo, valor in dados:
        ws[f"A{r}"] = rotulo; ws[f"A{r}"].font = Font(bold=True, color="FFFFFF"); ws[f"A{r}"].fill = navy; ws[f"A{r}"].border = border
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6); ws[f"B{r}"] = valor
        for c in range(2,7): ws.cell(r,c).border = border
        r += 1
    r += 2
    headers = ["PAVIMENTO", "DESCRIÇÃO", "ÁREA (m²)", "P.D. (m)", "VALOR MENSAL", "VALOR DO PERÍODO"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h); cell.fill = red; cell.font = Font(color="FFFFFF", bold=True); cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border
    r += 1
    total_mensal = total_periodo = 0.0
    for teto in obra.get("tetos", []):
        for idx, sv in enumerate(teto.get("servicos", [])):
            vals=[teto.get("nome", "") if idx == 0 else "", sv.get("descricao", ""), teto.get("area", 0) if idx == 0 else "", teto.get("pe_direito", 0) if idx == 0 else "", sv.get("total_mensal", 0), sv.get("total_periodo", 0)]
            for c,v in enumerate(vals,1): ws.cell(r,c,v); ws.cell(r,c).border=border; ws.cell(r,c).alignment=Alignment(vertical="center",wrap_text=True)
            ws.cell(r,5).number_format='R$ #,##0.00'; ws.cell(r,6).number_format='R$ #,##0.00'; r += 1
        ws.cell(r,2,f"SUBTOTAL — {teto.get('nome','')}"); ws.cell(r,5,teto.get("valor_mensal",0)); ws.cell(r,6,teto.get("valor_total",0))
        for c in range(1,7): ws.cell(r,c).fill=light; ws.cell(r,c).border=border; ws.cell(r,c).font=Font(bold=True)
        ws.cell(r,5).number_format='R$ #,##0.00'; ws.cell(r,6).number_format='R$ #,##0.00'
        total_mensal += teto.get("valor_mensal",0); total_periodo += teto.get("valor_total",0); r += 2
    ws.cell(r,4,"TOTAL GERAL"); ws.cell(r,5,total_mensal); ws.cell(r,6,total_periodo)
    for c in range(4,7): ws.cell(r,c).fill=red; ws.cell(r,c).font=Font(color="FFFFFF",bold=True,size=12); ws.cell(r,c).border=border
    ws.cell(r,5).number_format='R$ #,##0.00'; ws.cell(r,6).number_format='R$ #,##0.00'
    ws.freeze_panes="A15"; ws.sheet_view.showGridLines=False; ws.page_setup.orientation="landscape"; ws.page_setup.fitToWidth=1; ws.sheet_properties.pageSetUpPr.fitToPage=True
    wt=wb.create_sheet("Orçamento Técnico"); tech=linhas_tecnicas(obra)
    if tech:
        hs=list(tech[0])
        for c,h in enumerate(hs,1): wt.cell(1,c,h); wt.cell(1,c).fill=navy; wt.cell(1,c).font=Font(color="FFFFFF",bold=True); wt.cell(1,c).border=border
        for ri,row in enumerate(tech,2):
            for c,h in enumerate(hs,1): wt.cell(ri,c,row[h]); wt.cell(ri,c).border=border
        for c in range(1,len(hs)+1): wt.column_dimensions[get_column_letter(c)].width=18
        wt.column_dimensions["B"].width=38; wt.freeze_panes="A2"; wt.auto_filter.ref=wt.dimensions
        for row in wt.iter_rows(min_row=2):
            for c in [9,10,11,13,14]: row[c-1].number_format='R$ #,##0.00'
    wm=wb.create_sheet("Memória de Cálculo"); mr=1
    for teto in obra.get("tetos",[]):
        wm.cell(mr,1,teto.get("nome","")); wm.cell(mr,1).fill=red; wm.cell(mr,1).font=Font(color="FFFFFF",bold=True); wm.merge_cells(start_row=mr,start_column=1,end_row=mr,end_column=8); mr+=1
        hs=["CÓDIGO","DESCRIÇÃO","QTD.TOTAL","PESO UNIT.","PESO TOTAL","VALOR UNIT.","VALOR TOTAL","ÍNDICE"]
        for c,h in enumerate(hs,1): wm.cell(mr,c,h); wm.cell(mr,c).fill=gray; wm.cell(mr,c).font=Font(bold=True); wm.cell(mr,c).border=border
        mr+=1
        for m in teto.get("materiais",[]):
            vals=[m.get("codigo",""),m.get("descricao",""),m.get("quantidade",0),m.get("peso_unitario",0),m.get("peso_total",0),m.get("valor_unitario",0),m.get("valor_total",0),teto.get("indice_nome","")]
            for c,v in enumerate(vals,1): wm.cell(mr,c,v); wm.cell(mr,c).border=border
            wm.cell(mr,6).number_format='R$ #,##0.00'; wm.cell(mr,7).number_format='R$ #,##0.00'; mr+=1
        mr+=2
    for c,w in enumerate([14,42,14,14,14,14,16,25],1): wm.column_dimensions[get_column_letter(c)].width=w
    output=BytesIO(); wb.save(output); return output.getvalue()

def gerar_pdf_proposta(obra):
    if not REPORTLAB_DISPONIVEL:
        raise RuntimeError("A biblioteca reportlab não está instalada. Execute instalar_bibliotecas.bat.")
    output=BytesIO()
    doc=SimpleDocTemplate(output,pagesize=landscape(A4),rightMargin=12*mm,leftMargin=12*mm,topMargin=10*mm,bottomMargin=10*mm)
    styles=getSampleStyleSheet(); title=ParagraphStyle("TitleJ",parent=styles["Title"],fontName="Helvetica-Bold",fontSize=18,textColor=colors.HexColor("#182733"),alignment=1,spaceAfter=8); small=ParagraphStyle("SmallJ",parent=styles["Normal"],fontSize=8,leading=10)
    story=[]
    logo=Table([[Paragraph("<b>ANDAIMES<br/><font size='18'>JIRAU</font><br/>ESCORAMENTO</b>",ParagraphStyle("logo",parent=small,textColor=colors.white,alignment=1))]],colWidths=[48*mm],rowHeights=[28*mm]); logo.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#C91F2B")),("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    story += [logo,Spacer(1,4*mm),Paragraph("PROPOSTA COMERCIAL",title)]
    info=[["DATA",date.today().strftime("%d/%m/%Y"),"CLIENTE",obra.get("cliente","")],["OBRA",obra.get("nome",""),"REFERÊNCIA",obra.get("referencia","")],["ENDEREÇO",obra.get("endereco",""),"",""]]
    ti=Table(info,colWidths=[24*mm,92*mm,28*mm,116*mm]); ti.setStyle(TableStyle([("GRID",(0,0),(-1,-1),.5,colors.grey),("BACKGROUND",(0,0),(0,-1),colors.HexColor("#182733")),("BACKGROUND",(2,0),(2,-1),colors.HexColor("#182733")),("TEXTCOLOR",(0,0),(0,-1),colors.white),("TEXTCOLOR",(2,0),(2,-1),colors.white),("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    story += [ti,Spacer(1,5*mm)]
    data=[["PAVIMENTO","DESCRIÇÃO","ÁREA m²","P.D. m","VALOR MENSAL","VALOR DO PERÍODO"]]
    for teto in obra.get("tetos",[]):
        first=True
        for sv in teto.get("servicos",[]): data.append([teto.get("nome","") if first else "",sv.get("descricao",""),numero_br(teto.get("area",0)) if first else "",numero_br(teto.get("pe_direito",0)) if first else "",moeda(sv.get("total_mensal",0)),moeda(sv.get("total_periodo",0))]); first=False
        data.append(["",f"SUBTOTAL — {teto.get('nome','')}","","",moeda(teto.get("valor_mensal",0)),moeda(teto.get("valor_total",0))])
    data.append(["","TOTAL GERAL","","",moeda(sum(t.get("valor_mensal",0) for t in obra.get("tetos",[]))),moeda(sum(t.get("valor_total",0) for t in obra.get("tetos",[])))])
    tbl=Table(data,colWidths=[34*mm,85*mm,25*mm,22*mm,40*mm,44*mm],repeatRows=1)
    style=[("GRID",(0,0),(-1,-1),.5,colors.HexColor("#333333")),("BACKGROUND",(0,0),(-1,0),colors.HexColor("#C91F2B")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),7.5),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("ALIGN",(2,1),(-1,-1),"RIGHT")]
    row=1
    for teto in obra.get("tetos",[]): row += len(teto.get("servicos",[])); style += [("BACKGROUND",(0,row),(-1,row),colors.HexColor("#E9EEF2")),("FONTNAME",(0,row),(-1,row),"Helvetica-Bold")]; row += 1
    style += [("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#C91F2B")),("TEXTCOLOR",(0,-1),(-1,-1),colors.white),("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold")]
    tbl.setStyle(TableStyle(style)); story.append(tbl); doc.build(story); return output.getvalue()

LOGO_FILE = BASE_DIR / "assets" / "logo_jirau_oficial.svg"
if not LOGO_FILE.exists():
    LOGO_FILE = BASE_DIR / "logo_jirau.svg"
LOGO_URI = svg_data_uri(LOGO_FILE)

st.markdown(premium_css(), unsafe_allow_html=True)

with st.sidebar:
    st.markdown(f"""
    <div class="jirau-brand-box">
      <img src="{LOGO_URI}" alt="Jirau Engenharia">
      <div class="brand-sub">Sistema Inteligente de Orçamentos</div>
      <div class="jirau-version">ENTERPRISE V12</div>
    </div>
    """, unsafe_allow_html=True)
    paginas = [
        "🏠 Visão Geral",
        "1️⃣ Materiais",
        "2️⃣ Índices",
        "3️⃣ Obras",
        "4️⃣ Tetos",
        "5️⃣ Orçamento",
        "⚙️ Administração",
    ]
    pagina = st.radio("Navegação", paginas, label_visibility="collapsed")
    st.divider()
    progresso = sum([
        bool(db["materiais"]), bool(db["indices"]), bool(db["obras"]),
        any(o.get("tetos") for o in db["obras"]),
    ]) / 4
    st.caption("CONFIGURAÇÃO DO SISTEMA")
    st.progress(progresso)
    c1, c2, c3 = st.columns(3)
    c1.metric("MAT.", len(db["materiais"]))
    c2.metric("ÍND.", len(db["indices"]))
    c3.metric("OBRAS", len(db["obras"]))

st.markdown(f"""
<div class="jirau-header">
  <div class="jirau-header-inner">
    <div class="jirau-header-logo"><img src="{LOGO_URI}" alt="Jirau"></div>
    <div>
      <div class="jirau-eyebrow">Jirau Engenharia</div>
      <div class="jirau-title">Jirau Enterprise</div>
      <div class="jirau-subtitle">Plataforma profissional para orçamento de andaimes e escoramentos • V12</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

if pagina == "🏠 Visão Geral":
    st.markdown("""<div class="premium-hero"><h1>Inteligência aplicada ao orçamento.</h1><p>Centralize materiais, índices técnicos, obras, quantitativos e propostas em uma experiência exclusiva da Jirau Engenharia.</p></div>""", unsafe_allow_html=True)
    st.subheader("Visão geral da operação")
    st.caption("Fluxo estruturado para criar orçamentos técnicos com precisão, rastreabilidade e apresentação profissional.")
    cols = st.columns(5)
    etapas = [
        ("1", "Materiais", "Importe ou cadastre peças e preços."),
        ("2", "Índices", "Crie quantos sistemas construtivos precisar."),
        ("3", "Obras", "Cadastre cliente, referência e endereço."),
        ("4", "Tetos", "Escolha um índice diferente em cada teto."),
        ("5", "Orçamento", "Veja custos, pesos e quantitativos."),
    ]
    for col, (n, titulo, texto) in zip(cols, etapas):
        col.markdown(f'<div class="card"><b style="font-size:22px;color:#C7192D">{n}</b><h4>{titulo}</h4><span>{texto}</span></div>', unsafe_allow_html=True)
    st.divider()
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Materiais", len(db["materiais"]))
    m2.metric("Índices", len(db["indices"]))
    m3.metric("Obras", len(db["obras"]))
    m4.metric("Tetos", sum(len(o.get("tetos", [])) for o in db["obras"]))

elif pagina == "1️⃣ Materiais":
    st.subheader("1. Materiais")
    tab1, tab2, tab3 = st.tabs(["Importar Excel", "Cadastrar manualmente", "Lista de materiais"])

    with tab1:
        arquivo = st.file_uploader("Selecione a planilha Excel", type=["xlsx", "xls"])
        substituir = st.checkbox("Substituir a lista atual ao importar", value=True)
        if arquivo:
            try:
                novos, aba = detectar_materiais_excel(arquivo)
                st.success(f"{len(novos)} materiais encontrados na aba {aba}.")
                st.dataframe(pd.DataFrame(novos)[["codigo", "descricao", "peso", "valor"]].head(50), use_container_width=True, hide_index=True)
                if st.button("Confirmar importação", type="primary", use_container_width=True):
                    if substituir:
                        substituir_materiais_preservando_vinculos(db, novos)
                    else:
                        existentes = {m["codigo"] for m in db["materiais"]}
                        db["materiais"].extend(m for m in novos if m["codigo"] not in existentes)
                    salvar_db(db)
                    st.success("Materiais importados.")
                    st.rerun()
            except Exception as e:
                st.error(f"Erro ao importar: {e}")

    with tab2:
        with st.form("novo_material"):
            c1, c2 = st.columns([1, 3])
            codigo = c1.text_input("Código")
            descricao = c2.text_input("Descrição")
            c3, c4 = st.columns(2)
            peso = c3.number_input("Peso unitário (kg)", min_value=0.0, step=0.001, format="%.3f")
            valor = c4.number_input("Valor de locação", min_value=0.0, step=0.01)
            criar = st.form_submit_button("Adicionar material", type="primary", use_container_width=True)
            if criar and codigo.strip() and descricao.strip():
                db["materiais"].append({"id": uid(), "codigo": codigo.strip(), "descricao": descricao.strip(), "peso": peso, "valor": valor})
                salvar_db(db)
                st.rerun()

    with tab3:
        if db["materiais"]:
            busca = st.text_input("Pesquisar material")
            filtrados = [m for m in db["materiais"] if busca.lower() in f"{m['codigo']} {m['descricao']}".lower()]
            st.dataframe(pd.DataFrame(filtrados).rename(columns={"codigo":"Código","descricao":"Descrição","peso":"Peso (kg)","valor":"Locação (R$)"})[["Código","Descrição","Peso (kg)","Locação (R$)"]], use_container_width=True, hide_index=True)
        else:
            st.info("Nenhum material cadastrado.")

elif pagina == "2️⃣ Índices":
    st.subheader("2. Índices técnicos de escoramento")
    if not db["materiais"]:
        st.warning("Cadastre ou importe materiais primeiro.")
    else:
        with st.expander("➕ Criar novo índice", expanded=not db["indices"]):
            with st.form("novo_indice"):
                nome = st.text_input("Nome do índice", placeholder="Ex.: Teto padrão até 3,20 m")
                descricao = st.text_input("Descrição")
                criar = st.form_submit_button("Criar índice", type="primary", use_container_width=True)
                if criar and nome.strip():
                    db["indices"].append({"id": uid(), "nome": nome.strip(), "descricao": descricao.strip(), "area_base": 1.0, "pe_direito_base": 1.0, "itens": []})
                    salvar_db(db)
                    st.rerun()

        st.markdown("### Índice oficial da planilha de referência")
        st.caption("Carrega automaticamente materiais, quantidades e bases da aba TIPO oficial: 403,23 m² e P.D. 3,30 m.")
        if st.button("📥 Carregar / restaurar TIPO OFICIAL JIRAU 2026", type="primary", use_container_width=True):
            oficial = carregar_indice_tipo_oficial(db)
            salvar_db(db)
            st.session_state["indice_oficial_carregado"] = oficial["nome"]
            st.rerun()
        nome_oficial = st.session_state.pop("indice_oficial_carregado", None)
        if nome_oficial:
            st.success(f'Índice "{nome_oficial}" carregado com as bases e quantidades oficiais.')

        if db["indices"]:
            nomes = {f"{i['nome']} ({len(i.get('itens', []))} itens)": i for i in db["indices"]}
            escolha = st.selectbox("Índice para editar", list(nomes.keys()))
            indice = nomes[escolha]
            alertas_indice = auditar_indice(indice)
            for alerta in alertas_indice:
                st.error(f"⚠️ {alerta} Corrija as bases ou carregue o índice TIPO oficial antes de gerar a proposta.")

            c1, c2 = st.columns([3, 1])
            with c1:
                novo_nome = st.text_input("Nome", value=indice["nome"], key=f"nome_{indice['id']}")
                nova_desc = st.text_input("Descrição", value=indice.get("descricao", ""), key=f"desc_{indice['id']}")
                ba, bp, bv = st.columns(3)
                area_base = ba.number_input("Área-base do índice (m²)", min_value=0.01, value=float(indice.get("area_base", 1.0) or 1.0), step=0.01, key=f"area_base_{indice['id']}")
                pe_base = bp.number_input("Pé-direito-base (m)", min_value=0.01, value=float(indice.get("pe_direito_base", 1.0) or 1.0), step=0.01, key=f"pe_base_{indice['id']}")
                bv.metric("Volume-base", f"{area_base * pe_base:,.2f} m³".replace(",", "X").replace(".", ",").replace("X", "."))
                st.caption("Use a mesma área-base para todas as colunas. O volume-base é calculado por Área × Pé-direito.")
                if st.button("Salvar identificação e bases do índice", use_container_width=True):
                    indice["nome"] = novo_nome.strip() or indice["nome"]
                    indice["descricao"] = nova_desc.strip()
                    indice["area_base"] = float(area_base)
                    indice["pe_direito_base"] = float(pe_base)
                    recalcular_todos(db)
                    st.rerun()
            with c2:
                if st.button("Duplicar índice", use_container_width=True):
                    copia = deepcopy(indice)
                    copia["id"] = uid()
                    copia["nome"] = f"{indice['nome']} - Cópia"
                    db["indices"].append(copia)
                    salvar_db(db)
                    st.rerun()

            opcoes = {f"{m['codigo']} — {m['descricao']}": m for m in db["materiais"]}
            with st.expander("➕ Adicionar material ao índice", expanded=not indice.get("itens")):
                with st.form(f"item_{indice['id']}"):
                    label = st.selectbox("Material", list(opcoes.keys()))
                    st.caption("Informe as quantidades do projeto-base. A área é única para todas as colunas; os grupos verticais usarão também o volume-base.")
                    c1, c2, c3, c4 = st.columns(4)
                    vigas_h = c1.number_input("ESCORAMENTO DE VIGAS — HORIZONTAL", min_value=0.0, step=0.001, format="%.3f")
                    vigas_v = c2.number_input("ESCORAMENTO DE VIGAS — VERTICAL", min_value=0.0, step=0.001, format="%.3f")
                    lajes_h = c3.number_input("ESCORAMENTO DE LAJES — HORIZONTAL", min_value=0.0, step=0.001, format="%.3f")
                    lajes_v = c4.number_input("ESCORAMENTO DE LAJES — VERTICAL", min_value=0.0, step=0.001, format="%.3f")
                    c5, c6, c7 = st.columns(3)
                    reesc = c5.number_input("REESCORAMENTO DE VIGAS E LAJES", min_value=0.0, step=0.001, format="%.3f")
                    trav_pil = c6.number_input("TRAVAMENTO DE PILARES", min_value=0.0, step=0.001, format="%.3f")
                    trav_vig = c7.number_input("TRAVAMENTO DE VIGAS", min_value=0.0, step=0.001, format="%.3f")
                    adicionar = st.form_submit_button("Adicionar ou atualizar material", type="primary", use_container_width=True)
                    if adicionar:
                        mat = opcoes[label]
                        item = next((x for x in indice["itens"] if x["material_id"] == mat["id"]), None)
                        valores = {
                            "vigas_h": vigas_h, "vigas_v": vigas_v, "lajes_h": lajes_h,
                            "lajes_v": lajes_v, "reesc": reesc, "trav_pil": trav_pil, "trav_vig": trav_vig,
                        }
                        if item:
                            item.update(valores)
                        else:
                            indice["itens"].append({"material_id": mat["id"], **valores})
                        recalcular_todos(db)
                        st.rerun()

            itens_por_material = {item.get("material_id"): item for item in indice.get("itens", [])}
            linhas = []
            ids_linhas = []
            for mat in db["materiais"]:
                item = itens_por_material.get(mat["id"], {})
                total = consumo_total_item(item)
                linha = {"CÓDIGO": mat["codigo"], "DESCRIÇÃO": mat["descricao"], "PESO(kg)": float(mat.get("peso", 0)), "VLR.LOCAÇÃO": float(mat.get("valor", 0))}
                for chave, rotulo in CAMPOS_INDICE:
                    linha[rotulo] = float(item.get(chave, 0) or 0)
                linha.update({"QTD.TOTAL": total, "PESO TOTAL": total * float(mat.get("peso", 0)), "VALOR": total * float(mat.get("valor", 0))})
                linhas.append(linha)
                ids_linhas.append(mat["id"])

            st.markdown("### Planilha de quantidades do índice-base")
            st.info("Cole as quantidades do projeto-base. A mesma área-base vale para todas as colunas. O sistema escala grupos horizontais pela área e grupos verticais pelo volume.")
            st.caption("Área: VIGAS H, LAJES H e TRAV. VIG. | Volume: VIGAS V, LAJES V, REESC. e TRAV. PIL.")

            chave_filtro = f"ocultar_vazios_{indice['id']}"
            ocultar_vazios = st.session_state.get(chave_filtro, False)

            if ocultar_vazios:
                pares = [
                    (linha, material_id)
                    for linha, material_id in zip(linhas, ids_linhas)
                    if any(numero_float(linha.get(rotulo, 0)) != 0 for _, rotulo in CAMPOS_INDICE)
                ]
                linhas_exibidas = [p[0] for p in pares]
                ids_exibidos = [p[1] for p in pares]
            else:
                linhas_exibidas = linhas
                ids_exibidos = ids_linhas

            vazias = sum(
                1 for linha in linhas
                if not any(numero_float(linha.get(rotulo, 0)) != 0 for _, rotulo in CAMPOS_INDICE)
            )
            preenchidas = len(linhas) - vazias
            m1, m2 = st.columns(2)
            m1.metric("Linhas preenchidas", preenchidas)
            m2.metric("Linhas vazias", vazias)

            st.markdown("""
            <div style="overflow-x:auto;margin:8px 0 4px 0;">
              <table style="width:100%;min-width:1500px;border-collapse:collapse;text-align:center;font-size:12px;">
                <tr>
                  <th rowspan="2" style="border:1px solid #5d6d78;background:#9dc3e6;padding:8px;">CÓDIGO</th>
                  <th rowspan="2" style="border:1px solid #5d6d78;background:#9dc3e6;padding:8px;">DESCRIÇÃO</th>
                  <th rowspan="2" style="border:1px solid #5d6d78;background:#9dc3e6;padding:8px;">PESO(kg)</th>
                  <th rowspan="2" style="border:1px solid #5d6d78;background:#9dc3e6;padding:8px;">VLR.LOCAÇÃO</th>
                  <th rowspan="2" style="border:1px solid #5d6d78;background:#9dc3e6;padding:8px;">QTD.TOTAL</th>
                  <th colspan="2" style="border:1px solid #5d6d78;background:#5482a5;color:white;padding:8px;">ESCORAMENTO DE VIGAS</th>
                  <th colspan="2" style="border:1px solid #5d6d78;background:#5482a5;color:white;padding:8px;">ESCORAMENTO DE LAJES</th>
                  <th rowspan="2" style="border:1px solid #5d6d78;background:#9dc3e6;padding:8px;">REESCORAMENTO<br>DE VIGAS E LAJES</th>
                  <th rowspan="2" style="border:1px solid #5d6d78;background:#9dc3e6;padding:8px;">TRAVAMENTO<br>DE PILARES</th>
                  <th rowspan="2" style="border:1px solid #5d6d78;background:#9dc3e6;padding:8px;">TRAVAMENTO<br>DE VIGAS</th>
                </tr>
                <tr>
                  <th style="border:1px solid #5d6d78;background:#9dc3e6;padding:8px;">HORIZONTAL</th>
                  <th style="border:1px solid #5d6d78;background:#9dc3e6;padding:8px;">VERTICAL</th>
                  <th style="border:1px solid #5d6d78;background:#9dc3e6;padding:8px;">HORIZONTAL</th>
                  <th style="border:1px solid #5d6d78;background:#9dc3e6;padding:8px;">VERTICAL</th>
                </tr>
              </table>
            </div>
            """, unsafe_allow_html=True)

            editado = st.data_editor(
                pd.DataFrame(linhas_exibidas), use_container_width=True, hide_index=True, height=620,
                disabled=["CÓDIGO", "DESCRIÇÃO", "PESO(kg)", "VLR.LOCAÇÃO", "QTD.TOTAL", "PESO TOTAL", "VALOR"],
                column_config={rotulo: st.column_config.NumberColumn(rotulo, min_value=0.0, step=0.001, format="%.3f") for _, rotulo in CAMPOS_INDICE},
                key=f"editor_indice_{indice['id']}_{'limpo' if ocultar_vazios else 'completo'}",
            )

            def salvar_editor_indice():
                mapa_existente = {item.get("material_id"): item.copy() for item in indice.get("itens", [])}
                for pos, material_id in enumerate(ids_exibidos):
                    valores = {}
                    for chave, rotulo in CAMPOS_INDICE:
                        celula = editado.iloc[pos][rotulo]
                        valores[chave] = 0.0 if pd.isna(celula) else max(0.0, numero_float(celula))
                    if any(v != 0 for v in valores.values()):
                        mapa_existente[material_id] = {"material_id": material_id, **valores}
                    else:
                        mapa_existente.pop(material_id, None)
                indice["itens"] = list(mapa_existente.values())
                recalcular_todos(db)
                salvar_db(db)

            # Painel único de ações da planilha: evita vários botões repetidos e soltos.
            st.markdown("### Ações da planilha")
            with st.container(border=True):
                st.markdown("#### Edição e visualização")
                st.caption(
                    "Salve as alterações, esconda as linhas sem uso ou restaure a lista completa de materiais."
                )

                acao_salvar, acao_visualizacao = st.columns([1.35, 1])
                if acao_salvar.button(
                    "💾 Salvar alterações",
                    type="primary",
                    use_container_width=True,
                    key=f"salvar_indice_{indice['id']}",
                ):
                    salvar_editor_indice()
                    st.session_state["indice_salvo"] = True
                    st.rerun()

                if not ocultar_vazios:
                    if acao_visualizacao.button(
                        f"🧹 Ocultar {vazias} linhas vazias",
                        use_container_width=True,
                        disabled=vazias == 0,
                        key=f"ocultar_vazias_{indice['id']}",
                    ):
                        salvar_editor_indice()
                        st.session_state[chave_filtro] = True
                        st.session_state["linhas_limpas"] = vazias
                        st.rerun()
                else:
                    if acao_visualizacao.button(
                        "↩️ Mostrar lista completa",
                        use_container_width=True,
                        key=f"mostrar_todas_{indice['id']}",
                    ):
                        st.session_state[chave_filtro] = False
                        st.rerun()

                st.divider()
                st.markdown("#### Limpeza de valores")
                st.caption(
                    "Escolha uma coluna e apague todos os valores dela de uma só vez. "
                    "Código, descrição, peso e valor de locação não são alterados."
                )

                def limpar_coluna_indice(chave_coluna):
                    salvar_editor_indice()
                    for item in indice.get("itens", []):
                        item[chave_coluna] = 0.0
                    indice["itens"] = [
                        item for item in indice.get("itens", [])
                        if any(numero_float(item.get(chave, 0)) != 0 for chave, _ in CAMPOS_INDICE)
                    ]
                    recalcular_todos(db)
                    salvar_db(db)

                opcoes_limpeza = {
                    "Escoramento de vigas — Horizontal": "vigas_h",
                    "Escoramento de vigas — Vertical": "vigas_v",
                    "Escoramento de lajes — Horizontal": "lajes_h",
                    "Escoramento de lajes — Vertical": "lajes_v",
                    "Reescoramento de vigas e lajes": "reesc",
                    "Travamento de pilares": "trav_pil",
                    "Travamento de vigas": "trav_vig",
                }
                limpar_selecao, limpar_botao = st.columns([2.2, 1])
                coluna_escolhida = limpar_selecao.selectbox(
                    "Coluna que será zerada",
                    list(opcoes_limpeza.keys()),
                    key=f"coluna_limpeza_{indice['id']}",
                )
                limpar_botao.write("")
                limpar_botao.write("")
                if limpar_botao.button(
                    "🗑️ Limpar coluna",
                    use_container_width=True,
                    key=f"limpar_coluna_selecionada_{indice['id']}",
                ):
                    limpar_coluna_indice(opcoes_limpeza[coluna_escolhida])
                    st.session_state["coluna_limpa"] = coluna_escolhida
                    st.rerun()

                with st.expander("⚠️ Opções avançadas de limpeza"):
                    st.caption(
                        "Use esta opção apenas para reiniciar completamente as quantidades deste índice."
                    )
                    confirmar_limpeza_total = st.checkbox(
                        "Confirmo que desejo apagar todas as quantidades",
                        key=f"confirmar_limpeza_total_{indice['id']}",
                    )
                    if st.button(
                        "Apagar todas as quantidades do índice",
                        use_container_width=True,
                        disabled=not confirmar_limpeza_total,
                        key=f"limpar_tudo_{indice['id']}",
                    ):
                        salvar_editor_indice()
                        indice["itens"] = []
                        recalcular_todos(db)
                        salvar_db(db)
                        st.session_state["coluna_limpa"] = "Todas as colunas de quantidade"
                        st.rerun()

            coluna_limpa = st.session_state.pop("coluna_limpa", None)
            if coluna_limpa:
                st.success(f"{coluna_limpa} foi limpa com sucesso.")

            if st.session_state.pop("indice_salvo", False):
                st.success("Planilha do índice salva com sucesso.")
            removidas = st.session_state.pop("linhas_limpas", None)
            if removidas is not None:
                st.success(f"{removidas} linhas com apenas 0, vazio ou None foram removidas da visualização.")

            # Memória de cálculo no mesmo conceito da planilha Jirau.
            materiais_por_id = {m["id"]: m for m in db["materiais"]}
            qtd_grupo = {chave: 0.0 for chave, _ in CAMPOS_INDICE}
            peso_grupo = {chave: 0.0 for chave, _ in CAMPOS_INDICE}
            valor_grupo = {chave: 0.0 for chave, _ in CAMPOS_INDICE}

            for pos, material_id in enumerate(ids_exibidos):
                mat = materiais_por_id.get(material_id, {})
                peso_unit = numero_float(mat.get("peso", 0))
                valor_unit = numero_float(mat.get("valor", 0))
                for chave, rotulo in CAMPOS_INDICE:
                    celula = editado.iloc[pos][rotulo]
                    qtd = 0.0 if pd.isna(celula) else max(0.0, numero_float(celula))
                    qtd_grupo[chave] += qtd
                    peso_grupo[chave] += qtd * peso_unit
                    valor_grupo[chave] += qtd * valor_unit

            area_memoria = numero_float(area_base)
            volume_memoria = numero_float(area_base) * numero_float(pe_base)
            total_qtd = sum(qtd_grupo.values())
            total_peso = sum(peso_grupo.values())
            total_valor = sum(valor_grupo.values())

            st.markdown("### Memória de cálculo do índice")
            st.caption(
                "As quantidades informadas são do projeto-base. O sistema calcula automaticamente "
                "os índices unitários usando uma única área e um único pé-direito."
            )

            colunas_memoria = ["QTD.TOTAL"] + [rotulo for _, rotulo in CAMPOS_INDICE]
            linhas_memoria = []

            def linha_memoria(nome, total, valores):
                linha = {"INDICADOR": nome, "QTD.TOTAL": total}
                for (chave, rotulo), valor in zip(CAMPOS_INDICE, valores):
                    linha[rotulo] = valor
                return linha

            linhas_memoria.append(linha_memoria(
                "LOCAÇÃO MENSAL (R$)", total_valor,
                [valor_grupo[chave] for chave, _ in CAMPOS_INDICE]
            ))
            linhas_memoria.append(linha_memoria(
                "LOCAÇÃO DIÁRIA (R$)", total_valor / 30,
                [valor_grupo[chave] / 30 for chave, _ in CAMPOS_INDICE]
            ))
            linhas_memoria.append(linha_memoria(
                "PESO TOTAL (kg)", total_peso,
                [peso_grupo[chave] for chave, _ in CAMPOS_INDICE]
            ))
            peso_horizontal = sum(peso_grupo[chave] for chave in CAMPOS_AREA)
            peso_vertical = sum(peso_grupo[chave] for chave in CAMPOS_VOLUME)
            valor_horizontal = sum(valor_grupo[chave] for chave in CAMPOS_AREA)
            valor_vertical = sum(valor_grupo[chave] for chave in CAMPOS_VOLUME)
            linhas_memoria.append(linha_memoria(
                "kg/m² (somente horizontais)", peso_horizontal / area_memoria if area_memoria > 0 else 0,
                [peso_grupo[chave] / area_memoria if chave in CAMPOS_AREA and area_memoria > 0 else 0 for chave, _ in CAMPOS_INDICE]
            ))
            linhas_memoria.append(linha_memoria(
                "kg/m³ (somente verticais)", peso_vertical / volume_memoria if volume_memoria > 0 else 0,
                [peso_grupo[chave] / volume_memoria if chave in CAMPOS_VOLUME and volume_memoria > 0 else 0 for chave, _ in CAMPOS_INDICE]
            ))
            linhas_memoria.append(linha_memoria(
                "PREÇO/m² (somente horizontais)", valor_horizontal / area_memoria if area_memoria > 0 else 0,
                [valor_grupo[chave] / area_memoria if chave in CAMPOS_AREA and area_memoria > 0 else 0 for chave, _ in CAMPOS_INDICE]
            ))
            linhas_memoria.append(linha_memoria(
                "PREÇO/m³ (somente verticais)", valor_vertical / volume_memoria if volume_memoria > 0 else 0,
                [valor_grupo[chave] / volume_memoria if chave in CAMPOS_VOLUME and volume_memoria > 0 else 0 for chave, _ in CAMPOS_INDICE]
            ))

            memoria_df = pd.DataFrame(linhas_memoria)
            st.dataframe(
                memoria_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    coluna: st.column_config.NumberColumn(coluna, format="%.3f")
                    for coluna in colunas_memoria
                },
            )

            preco_m2_h = (
                valor_grupo["vigas_h"] + valor_grupo["lajes_h"]
            ) / area_memoria if area_memoria > 0 else 0
            preco_m3_v = (
                valor_grupo["vigas_v"] + valor_grupo["lajes_v"]
            ) / volume_memoria if volume_memoria > 0 else 0

            st.markdown("### Índices finais de vigas e lajes")
            i1, i2 = st.columns(2)
            i1.metric("PREÇO m² (H)", moeda(preco_m2_h))
            i1.caption("VIGAS H + LAJES H, divididos pela área-base.")
            i2.metric("PREÇO m³ (V)", moeda(preco_m3_v))
            i2.caption("VIGAS V + LAJES V, divididos pelo volume-base.")

            s1, s2, s3, s4 = st.columns(4)
            s1.metric("Quantidade total", f"{total_qtd:.3f}")
            s2.metric("Peso total", f"{total_peso:.3f} kg")
            s3.metric("Locação mensal", moeda(total_valor))
            s4.metric("Locação diária", moeda(total_valor / 30))

            st.markdown("### Índices complementares")
            complementares = pd.DataFrame([
                {
                    "GRUPO": rotulo,
                    "PREÇO/m²": valor_grupo[chave] / area_memoria if chave in CAMPOS_AREA and area_memoria > 0 else None,
                    "PREÇO/m³": valor_grupo[chave] / volume_memoria if chave in CAMPOS_VOLUME and volume_memoria > 0 else None,
                    "kg/m²": peso_grupo[chave] / area_memoria if chave in CAMPOS_AREA and area_memoria > 0 else None,
                    "kg/m³": peso_grupo[chave] / volume_memoria if chave in CAMPOS_VOLUME and volume_memoria > 0 else None,
                }
                for chave, rotulo in CAMPOS_INDICE
            ])
            st.dataframe(
                complementares,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "PREÇO/m²": st.column_config.NumberColumn("PREÇO/m²", format="R$ %.3f"),
                    "PREÇO/m³": st.column_config.NumberColumn("PREÇO/m³", format="R$ %.3f"),
                    "kg/m²": st.column_config.NumberColumn("kg/m²", format="%.3f"),
                    "kg/m³": st.column_config.NumberColumn("kg/m³", format="%.3f"),
                },
            )

elif pagina == "3️⃣ Obras":
    st.subheader("3. Obras")
    with st.form("nova_obra"):
        c1, c2 = st.columns(2)
        nome = c1.text_input("Nome da obra")
        cliente = c2.text_input("Cliente")
        c3, c4 = st.columns(2)
        referencia = c3.text_input("Referência/JLOC")
        endereco = c4.text_input("Endereço")
        criar = st.form_submit_button("Cadastrar obra", type="primary", use_container_width=True)
        if criar and nome.strip():
            db["obras"].append({"id": uid(), "nome": nome.strip(), "cliente": cliente.strip(), "referencia": referencia.strip(), "endereco": endereco.strip(), "tetos": []})
            salvar_db(db)
            st.rerun()
    if db["obras"]:
        st.dataframe(pd.DataFrame([{"Obra":o["nome"],"Cliente":o.get("cliente",""),"Referência":o.get("referencia",""),"Tetos":len(o.get("tetos",[]))} for o in db["obras"]]), use_container_width=True, hide_index=True)

elif pagina == "4️⃣ Tetos":
    st.subheader("4. Tetos, pavimentos e composição da proposta")
    st.caption("Cada pavimento usa um índice próprio e gera as quatro linhas comerciais da planilha Jirau.")
    if not db["obras"]:
        st.warning("Cadastre uma obra primeiro.")
    elif not db["indices"]:
        st.warning("Crie pelo menos um índice primeiro.")
    else:
        obra_id = st.selectbox("Obra", [o["id"] for o in db["obras"]], format_func=lambda oid: next(o["nome"] for o in db["obras"] if o["id"] == oid))
        obra = next(o for o in db["obras"] if o["id"] == obra_id)
        with st.container(border=True):
            c1, c2, c3 = st.columns(3)
            nome_teto = c1.text_input("Nome do teto/pavimento", placeholder="Ex.: 1º pavimento", key="novo_teto_nome")
            area = c2.number_input("Área (m²)", min_value=0.0, step=0.01, key="novo_teto_area")
            pe_direito = c3.number_input("Pé-direito (m)", min_value=0.0, step=0.01, key="novo_teto_pe_direito")
            c4, c5 = st.columns(2)
            dias = c4.number_input("Dias de locação", min_value=1, value=30, key="novo_teto_dias")
            indice_id = c5.selectbox("Índice a utilizar", [i["id"] for i in db["indices"]], format_func=lambda iid: next(i["nome"] for i in db["indices"] if i["id"] == iid), key="novo_teto_indice")
            if st.button("Adicionar teto", type="primary", use_container_width=True, key="botao_adicionar_teto"):
                erros=[]
                if not nome_teto.strip(): erros.append("informe o nome do teto/pavimento")
                if area <= 0: erros.append("informe uma área maior que zero")
                if pe_direito <= 0: erros.append("informe um pé-direito maior que zero")
                if erros:
                    st.error("Não foi possível adicionar: " + "; ".join(erros) + ".")
                else:
                    teto={"id":uid(),"nome":nome_teto.strip(),"area":float(area),"pe_direito":float(pe_direito),"dias":int(dias),"indice_id":indice_id}
                    recalcular_teto(db,teto)
                    obra.setdefault("tetos",[]).append(teto)
                    salvar_db(db)
                    st.session_state["teto_adicionado"]=teto["nome"]
                    st.rerun()
        msg=st.session_state.pop("teto_adicionado",None)
        if msg: st.success(f'Teto "{msg}" adicionado com sucesso.')

        if obra.get("tetos"):
            st.markdown("### Pavimentos cadastrados")
            st.dataframe(pd.DataFrame([{"Teto":t["nome"],"Área (m²)":t["area"],"Pé-direito (m)":t["pe_direito"],"Volume (m³)":t.get("volume",0),"Índice":t.get("indice_nome",""),"Dias":t["dias"],"Mensal":moeda(t.get("valor_mensal",0)),"Período":moeda(t.get("valor_total",0))} for t in obra["tetos"]]),use_container_width=True,hide_index=True)
            teto_sel=st.selectbox("Teto para configurar, alterar ou excluir", obra["tetos"], format_func=lambda t:t["nome"])
            st.markdown("### Configuração comercial do pavimento")
            st.caption("Percentual controla a parcela utilizada do sistema. Jogos multiplica a quantidade do conjunto. O campo FATOR da planilha é apenas informativo e não multiplica o preço.")
            config=teto_sel.get("servicos_config",{})
            linhas=[]
            for serv in SERVICOS_PROPOSTA:
                cfg=config.get(serv["chave"],{"percentual":100.0,"quantidade_jogos":1.0})
                linhas.append({"SERVIÇO":serv["descricao"],"JOGOS":cfg.get("quantidade_jogos",1.0),"PERCENTUAL (%)":cfg.get("percentual",100.0)})
            edit_cfg=st.data_editor(pd.DataFrame(linhas),use_container_width=True,hide_index=True,disabled=["SERVIÇO"],column_config={"JOGOS":st.column_config.NumberColumn(min_value=0.0,step=0.1,format="%.2f"),"PERCENTUAL (%)":st.column_config.NumberColumn(min_value=0.0,step=1.0,format="%.2f")},key=f"config_servicos_{teto_sel['id']}")
            c1,c2,c3=st.columns(3)
            novo_indice=c1.selectbox("Índice",db["indices"],index=max(0,next((n for n,i in enumerate(db["indices"]) if i["id"]==teto_sel.get("indice_id")),0)),format_func=lambda i:i["nome"],key=f"indice_teto_{teto_sel['id']}")
            novos_dias=c2.number_input("Dias",min_value=1,value=int(teto_sel.get("dias",30)),key=f"dias_teto_{teto_sel['id']}")
            novo_pe=c3.number_input("Pé-direito",min_value=0.01,value=float(teto_sel.get("pe_direito",0.01)),step=0.01,key=f"pe_teto_{teto_sel['id']}")
            b1,b2=st.columns([2,1])
            if b1.button("Salvar configuração e recalcular",type="primary",use_container_width=True,key=f"salvar_teto_{teto_sel['id']}"):
                novo_cfg={}
                for pos,serv in enumerate(SERVICOS_PROPOSTA):
                    row=edit_cfg.iloc[pos]
                    novo_cfg[serv["chave"]]={"quantidade_jogos":float(row["JOGOS"]),"percentual":float(row["PERCENTUAL (%)"]),"fator":1.0}
                teto_sel["servicos_config"]=novo_cfg
                teto_sel["indice_id"]=novo_indice["id"]
                teto_sel["dias"]=int(novos_dias)
                teto_sel["pe_direito"]=float(novo_pe)
                recalcular_teto(db,teto_sel)
                salvar_db(db)
                st.success("Configuração salva e orçamento recalculado.")
                st.rerun()
            if b2.button("Excluir teto",use_container_width=True,key=f"excluir_teto_{teto_sel['id']}"):
                obra["tetos"]=[t for t in obra["tetos"] if t["id"]!=teto_sel["id"]]
                salvar_db(db); st.rerun()

            recalcular_teto(db,teto_sel)
            st.markdown("### Prévia da proposta deste pavimento")
            prev=[]
            for sv in teto_sel.get("servicos",[]):
                prev.append({"QUANT. (JOGO)":sv["quantidade_jogos"],"DESCRIÇÃO":sv["descricao"],"%":sv["percentual"],"P.D. (m)":teto_sel["pe_direito"],"PAVIMENTO":teto_sel["nome"],"ÁREA m²":teto_sel["area"],"m³":teto_sel["volume"],"UNIT. R$/DIA":sv["unitario_dia"],"TOTAL R$/MÊS":sv["total_mensal"],"HORIZONTAL R$/m²":sv["horizontal_rs_m2"],"VERTICAL R$/m³":sv["vertical_rs_m3"]})
            st.dataframe(pd.DataFrame(prev),use_container_width=True,hide_index=True,column_config={c:st.column_config.NumberColumn(format="R$ %.2f") for c in ["UNIT. R$/DIA","TOTAL R$/MÊS","HORIZONTAL R$/m²","VERTICAL R$/m³"]})

elif pagina == "5️⃣ Orçamento":
    st.subheader("5. Proposta comercial e orçamento final")
    obras = [o for o in db["obras"] if o.get("tetos")]
    if not obras:
        st.warning("Ainda não há tetos lançados.")
    else:
        obra = st.selectbox("Obra", obras, format_func=lambda o: o["nome"])
        recalcular_todos(db)
        total = sum(t.get("valor_total", 0) for t in obra["tetos"])
        mensal = sum(t.get("valor_mensal", 0) for t in obra["tetos"])
        area_total = sum(t.get("area", 0) for t in obra["tetos"])
        volume_total = sum(t.get("volume", 0) for t in obra["tetos"])
        peso_total = sum(sum(m.get("peso_total", 0) for m in t.get("materiais", [])) for t in obra["tetos"])
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Área total", f"{numero_br(area_total)} m²")
        c2.metric("Volume total", f"{numero_br(volume_total)} m³")
        c3.metric("Peso total", f"{numero_br(peso_total)} kg")
        c4.metric("Locação mensal", moeda(mensal))
        c5.metric("Total do período", moeda(total))
        tab_comercial, tab_tecnico, tab_memoria = st.tabs(["📄 Proposta Comercial", "📐 Orçamento Técnico", "🧮 Memória de Cálculo"])
        with tab_comercial:
            st.markdown("### Documento que será entregue ao cliente")
            st.caption("Visual limpo, sem índices e sem lista de materiais. Cada pavimento aparece em um bloco e os subtotais ficam destacados.")
            st.markdown(f"""<div style='background:white;border:1px solid #E5E9F0;padding:24px;border-radius:18px;box-shadow:0 12px 30px rgba(23,32,42,.08)'><div style='display:flex;gap:22px;align-items:center;border-bottom:4px solid #C7192D;padding-bottom:15px'><div style='background:#fff;padding:8px 18px;border-right:1px solid #E5E9F0'><img src='{LOGO_URI}' style='width:130px;max-height:70px'></div><div><h2 style='margin:0;color:#182733'>PROPOSTA COMERCIAL</h2><div><b>Cliente:</b> {obra.get('cliente','')}</div><div><b>Obra:</b> {obra.get('nome','')}</div><div><b>Endereço:</b> {obra.get('endereco','')}</div><div><b>Referência:</b> {obra.get('referencia','')}</div></div></div></div>""", unsafe_allow_html=True)
            for teto in obra.get("tetos", []):
                st.markdown(f"#### {teto['nome']}")
                a, p, v, d = st.columns(4)
                a.metric("Área", f"{numero_br(teto.get('area',0))} m²")
                p.metric("Pé-direito", f"{numero_br(teto.get('pe_direito',0))} m")
                v.metric("Volume", f"{numero_br(teto.get('volume',0))} m³")
                d.metric("Prazo", f"{teto.get('dias',30)} dias")
                dados_serv = [{"DESCRIÇÃO": sv["descricao"], "VALOR MENSAL": sv["total_mensal"], "VALOR DO PERÍODO": sv["total_periodo"]} for sv in teto.get("servicos", [])]
                st.dataframe(pd.DataFrame(dados_serv), use_container_width=True, hide_index=True, column_config={"VALOR MENSAL": st.column_config.NumberColumn(format="R$ %.2f"), "VALOR DO PERÍODO": st.column_config.NumberColumn(format="R$ %.2f")})
                st.markdown(f"**Subtotal mensal:** {moeda(teto.get('valor_mensal',0))} &nbsp;&nbsp; | &nbsp;&nbsp; **Subtotal do período:** {moeda(teto.get('valor_total',0))}")
                st.divider()
            st.markdown(f"## TOTAL GERAL: {moeda(total)}")
            xlsx = gerar_excel_proposta(obra)
            d1, d2 = st.columns(2)
            d1.download_button("Baixar proposta profissional em Excel", xlsx, file_name=f"Proposta_Jirau_{obra['nome'].replace(' ','_')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            if REPORTLAB_DISPONIVEL:
                pdf = gerar_pdf_proposta(obra)
                d2.download_button("Baixar proposta comercial em PDF", pdf, file_name=f"Proposta_Jirau_{obra['nome'].replace(' ','_')}.pdf", mime="application/pdf", use_container_width=True)
            else:
                d2.warning("PDF indisponível: execute instalar_bibliotecas.bat e reinicie o aplicativo.")
        with tab_tecnico:
            st.markdown("### Quadro técnico no padrão da planilha Jirau")
            proposta_df = pd.DataFrame(linhas_tecnicas(obra))
            st.dataframe(proposta_df, use_container_width=True, hide_index=True, column_config={c: st.column_config.NumberColumn(format="R$ %.2f") for c in ["UNIT. R$/DIA", "TOTAL R$/MÊS", "TOTAL PERÍODO", "HORIZONTAL R$/m²", "VERTICAL R$/m³"]})
            st.caption("TOTAL R$/MÊS = Área × preço horizontal + Volume × preço vertical, considerando percentual e jogos. TOTAL PERÍODO = mensal × dias ÷ 30.")
        with tab_memoria:
            st.markdown("### Conferência interna de materiais e cálculos")
            teto = st.selectbox("Pavimento", obra["tetos"], format_func=lambda t: t["nome"], key="memoria_teto")
            if teto.get("materiais"):
                det = pd.DataFrame(teto["materiais"]).rename(columns={"codigo":"CÓDIGO","descricao":"DESCRIÇÃO","peso_unitario":"PESO(kg)","valor_unitario":"VLR.LOCAÇÃO","vigas_h":"ESCORAMENTO DE VIGAS — HORIZONTAL","vigas_v":"ESCORAMENTO DE VIGAS — VERTICAL","lajes_h":"ESCORAMENTO DE LAJES — HORIZONTAL","lajes_v":"ESCORAMENTO DE LAJES — VERTICAL","reesc":"REESCORAMENTO DE VIGAS E LAJES","trav_pil":"TRAVAMENTO DE PILARES","trav_vig":"TRAVAMENTO DE VIGAS","quantidade":"QTD.TOTAL","peso_total":"PESO TOTAL","valor_total":"VALOR"})
                ordem = ["CÓDIGO","DESCRIÇÃO","PESO(kg)","VLR.LOCAÇÃO","QTD.TOTAL","ESCORAMENTO DE VIGAS — HORIZONTAL","ESCORAMENTO DE VIGAS — VERTICAL","ESCORAMENTO DE LAJES — HORIZONTAL","ESCORAMENTO DE LAJES — VERTICAL","REESCORAMENTO DE VIGAS E LAJES","TRAVAMENTO DE PILARES","TRAVAMENTO DE VIGAS","PESO TOTAL","VALOR"]
                st.dataframe(det[[c for c in ordem if c in det.columns]], use_container_width=True, hide_index=True)

elif pagina == "⚙️ Administração":
    st.subheader("Administração")
    st.warning("As exclusões abaixo são permanentes.")

    if db["indices"]:
        indice = st.selectbox("Índice para excluir", db["indices"], format_func=lambda i:i["nome"])
        em_uso = sum(1 for o in db["obras"] for t in o.get("tetos",[]) if t.get("indice_id") == indice["id"])
        st.caption(f"Este índice está sendo usado em {em_uso} teto(s).")
        if st.button("Excluir índice", disabled=em_uso > 0):
            db["indices"] = [i for i in db["indices"] if i["id"] != indice["id"]]
            salvar_db(db)
            st.rerun()

    if db["obras"]:
        obra = st.selectbox("Obra para excluir", db["obras"], format_func=lambda o:o["nome"], key="adm_obra")
        if st.button("Excluir obra completa"):
            db["obras"] = [o for o in db["obras"] if o["id"] != obra["id"]]
            salvar_db(db)
            st.rerun()

    st.divider()
    backup = json.dumps(db, ensure_ascii=False, indent=2).encode("utf-8")
    st.download_button("Baixar backup dos dados", backup, file_name="jirau_db_backup.json", mime="application/json", use_container_width=True)


st.markdown("""<div class="jirau-footer">JIRAU ENTERPRISE V12 • Sistema Inteligente de Orçamentos • Uso exclusivo Jirau Engenharia</div>""", unsafe_allow_html=True)
